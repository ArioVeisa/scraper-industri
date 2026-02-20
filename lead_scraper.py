import re
import time
import requests
from bs4 import BeautifulSoup
from googlesearch import search
import json
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scraper_config import *

class LeadScraper:
    def __init__(self):
        self.results = []
        self.seen_urls = set()
        
    def extract_email(self, text):
        """Extract emails from text"""
        emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
        return list(set(emails))
    
    def extract_phone(self, text):
        """Extract Indonesian phone numbers"""
        # Format: 08xx, +62, (031), 031-, dll
        patterns = [
            r'\+62\s?\d{2,3}[\s-]?\d{3,4}[\s-]?\d{3,4}',
            r'0\d{2,3}[\s-]?\d{3,4}[\s-]?\d{3,4}',
            r'\(\d{3}\)\s?\d{3,4}[\s-]?\d{3,4}'
        ]
        phones = []
        for pattern in patterns:
            phones.extend(re.findall(pattern, text))
        return list(set(phones))
    
    def google_dork_search(self, query, location, num_results=5):
        """Search Google with dorking"""
        dork_query = f'{query} {location} (email OR kontak OR "hubungi kami" OR telepon)'
        print(f"[DORK] Searching: {dork_query}")
        
        try:
            results = []
            for url in search(dork_query, num_results=num_results, lang="id", sleep_interval=5):
                if url not in self.seen_urls:
                    self.seen_urls.add(url)
                    results.append(url)
                    print(f"  Found: {url}")
            return results
        except Exception as e:
            print(f"  Error: {e}")
            return []
    
    def crawl_website(self, url):
        """Crawl website untuk extract kontak"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract title
            title = soup.find('title')
            title = title.text.strip() if title else urlparse(url).netloc
            
            # Cari halaman contact
            contact_links = []
            for link in soup.find_all('a', href=True):
                href = link['href'].lower()
                if any(word in href for word in ['contact', 'kontak', 'about', 'tentang']):
                    full_url = urljoin(url, link['href'])
                    contact_links.append(full_url)
            
            # Extract dari halaman utama
            text = soup.get_text()
            emails = self.extract_email(text)
            phones = self.extract_phone(text)
            
            # Extract dari halaman contact (max 2 halaman)
            for contact_url in contact_links[:2]:
                try:
                    time.sleep(1)
                    resp = requests.get(contact_url, headers=headers, timeout=10)
                    contact_soup = BeautifulSoup(resp.text, 'html.parser')
                    contact_text = contact_soup.get_text()
                    emails.extend(self.extract_email(contact_text))
                    phones.extend(self.extract_phone(contact_text))
                except:
                    pass
            
            # Extract address (cari pattern alamat Indonesia)
            address = ""
            address_patterns = [
                r'Jl\.?\s+[A-Za-z0-9\s,.-]+(?:Surabaya|Sidoarjo|Gresik)',
                r'Alamat[:\s]+([^\n]{20,100})'
            ]
            for pattern in address_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    address = match.group(0).strip()
                    break
            
            return {
                'name': title,
                'url': url,
                'emails': list(set(emails)),
                'phones': list(set(phones)),
                'address': address
            }
            
        except Exception as e:
            print(f"  Error crawling {url}: {e}")
            return None
    
    def scrape_category(self, business_query, location, num_results=5):
        """Scrape satu kategori bisnis di satu lokasi"""
        print(f"\n{'='*60}")
        print(f"Target: {business_query} - {location}")
        print(f"{'='*60}")
        
        # Google Dork Search
        urls = self.google_dork_search(business_query, location, num_results=num_results)
        
        # Crawl each URL
        for url in urls:
            print(f"\n[CRAWL] {url}")
            data = self.crawl_website(url)
            if data and (data['emails'] or data['phones']):
                data['category'] = business_query
                data['location'] = location
                self.results.append(data)
                print(f"  ✓ Emails: {len(data['emails'])}, Phones: {len(data['phones'])}")
            time.sleep(2)  # Rate limiting
    
    def run_full_scrape(self, include_micro=True, results_per_category=5, max_categories=None):
        """Run full scraping campaign"""
        all_queries = BUSINESS_QUERIES.copy()
        if include_micro:
            all_queries.extend(MICRO_BUSINESS_QUERIES)
        
        # Limit categories if specified
        if max_categories:
            all_queries = all_queries[:max_categories]
        
        total_searches = len(all_queries) * len(LOCATIONS)
        print("Starting Lead Scraping Campaign...")
        print(f"Targets: {len(all_queries)} categories × {len(LOCATIONS)} locations = {total_searches} searches")
        print(f"Results per category: {results_per_category}")
        print(f"Estimated URLs to crawl: ~{total_searches * results_per_category}")
        print(f"Estimated time: ~{(total_searches * results_per_category * 7) // 60} minutes\n")
        
        for location in LOCATIONS:
            for query in all_queries:
                self.scrape_category(query, location, num_results=results_per_category)
        
        print(f"\n{'='*60}")
        print(f"SCRAPING COMPLETE: {len(self.results)} leads found")
        print(f"{'='*60}")
    
    def export_csv(self, filename="leads.csv"):
        """Export ke CSV (legacy support)"""
        import csv
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Nama', 'Kategori', 'Lokasi', 'Alamat', 'Email', 'No HP', 'Website'])
            
            for lead in self.results:
                emails = '; '.join(lead['emails'])
                phones = '; '.join(lead['phones'])
                writer.writerow([
                    lead['name'],
                    lead['category'],
                    lead['location'],
                    lead['address'],
                    emails,
                    phones,
                    lead['url']
                ])
        
        print(f"✓ Exported to {filename}")
    
    def export_excel(self, filename="erp_leads.xlsx"):
        """Export ke Excel dengan formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "ERP Leads"
        
        # Header styling
        headers = ['Nama', 'Kategori', 'Lokasi', 'Alamat', 'Email', 'No HP', 'Website']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_idx, lead in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=lead['name'])
            ws.cell(row=row_idx, column=2, value=lead['category'])
            ws.cell(row=row_idx, column=3, value=lead['location'])
            ws.cell(row=row_idx, column=4, value=lead['address'])
            ws.cell(row=row_idx, column=5, value='; '.join(lead['emails']))
            ws.cell(row=row_idx, column=6, value='; '.join(lead['phones']))
            ws.cell(row=row_idx, column=7, value=lead['url'])
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        wb.save(filename)
        print(f"✓ Exported to {filename}")
    
    def export_json(self, filename="leads.json"):
        """Export ke JSON"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)
        print(f"✓ Exported to {filename}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='ERP Lead Scraper')
    parser.add_argument('--locations', '-l', nargs='+', help='Custom locations (e.g., -l Surabaya Malang)')
    parser.add_argument('--no-micro', action='store_true', help='Skip micro business queries')
    parser.add_argument('--output', '-o', default='erp_leads.xlsx', help='Output filename')
    parser.add_argument('--results', '-r', type=int, default=5, help='Results per category (default: 5)')
    parser.add_argument('--max-categories', '-m', type=int, help='Limit number of categories to scrape')
    
    args = parser.parse_args()
    
    # Override locations if provided
    if args.locations:
        import scraper_config
        scraper_config.LOCATIONS = args.locations
        print(f"Custom locations: {', '.join(args.locations)}")
    
    scraper = LeadScraper()
    scraper.run_full_scrape(
        include_micro=not args.no_micro,
        results_per_category=args.results,
        max_categories=args.max_categories
    )
    scraper.export_excel(args.output)
    scraper.export_json(args.output.replace('.xlsx', '.json'))
